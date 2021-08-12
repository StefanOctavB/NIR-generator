from flask import Blueprint, render_template,request,flash,session, url_for, redirect, send_file, send_from_directory, abort, current_app
from flask.helpers import send_file
from flask.json import jsonify
from flask_login import login_required, current_user
from .models import Nir, Produs
from . import db
import json
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
import os
import requests
import time


views = Blueprint('views', __name__)

@views.route('/')
@login_required
def home():
    return render_template("home.html", user=current_user)

def getProductValues():
    produs = {
        "nume_produs":request.form.get('nume_produs'),
        "cod_produs": request.form.get('cod_produs'),
        "cantitate_produs":int(request.form.get('cantitate_produs')),
        "pret_achizitie_produs": float(request.form.get('pret_achizitie_produs')),
        "moneda_achizitie":request.form.get('moneda_achizitie'),
        "pret_vanzare_produs":float(request.form.get('pret_vanzare_produs')),
        "moneda_vanzare": request.form.get('moneda_vanzare')
    }
    return produs

def getNirValues():
    nir = {
        "nume_firma" : request.form.get('nume_firma'),
        "nume_furnizor" : request.form.get('nume_furnizor'),
        "numar_factura" : request.form.get('numar_factura'),
        "nume_membri" : request.form.get('nume_membri'),
        "nume_gestionar" : request.form.get('nume_gestionar'),
        "cost_transport" : float(request.form.get('cost_transport')),
        "moneda_transport" : request.form.get('moneda_transport'),
        "taxe_tva" : float(request.form.get('taxe_tva')),
        "moneda_taxe_tva" : request.form.get('moneda_taxe_tva'),
        "data_factura": request.form.get('data_factura')
    }
    return nir

headings = ("Nume produs", "Cod produs", "Cantitate", "Pret achizitie", "Pret vanzare","")
@views.route('/genereaza-nir', methods=['GET', 'POST'])
@login_required
def genereaza_nir():
    if request.method == "GET":
        if not session.get("data"):
            session["data"] = []
            return render_template("generateNIR.html", user=current_user, headings=headings, data=session.get("data"))
        else:
            return render_template("generateNIR.html", user=current_user, headings=headings, data=session.get("data"))
    elif request.method == 'POST':
        if request.form['btn'] == "adauga_produs":
            produs = getProductValues()
            if produs["nume_produs"] == " ":
                flash("Nume produs nu a fost introdus", category='error')
            elif produs["cod_produs"] == " ":
                flash("Cod produs nu a fost introdus", category='error')
            else:
                #return send_file ("/Users/octav/Desktop/work/OliBee/Gestiune/generateNir/website/static/nir-2.xlsx", as_attachment=True,mimetype='application/vnd.ms-excel',attachment_filename="nir-2.xlsx")
                flash("Produs adaugat", category='success')
                session.get("data").append(produs)
                return redirect(url_for('views.genereaza_nir'))
        elif request.form['btn'] == "genereaza_nir":
            nir = getNirValues()
            if not session.get('data'):
                flash("Nu se poate genera un NIR care nu contine produse", category='error')
                return redirect(url_for('views.genereaza_nir'))
            else:
                nir_nou = Nir(user_id=current_user.id,
                                    nume_firma=nir["nume_firma"], 
                                    nume_furnizor = nir["nume_furnizor"], 
                                    numar_factura=nir["numar_factura"],
                                    nume_membri=nir["nume_membri"],
                                    nume_gestionar=nir["nume_gestionar"],
                                    cost_transport=nir["cost_transport"],
                                    moneda_transport=nir["moneda_transport"],
                                    taxe_tva=nir["taxe_tva"],
                                    moneda_taxe_tva=nir["moneda_taxe_tva"],
                                    data_factura = nir["data_factura"])
                db.session.add(nir_nou)
                db.session.commit()
                for produs in session.get("data"):
                    produs_nou = Produs(nume_produs=produs["nume_produs"], 
                                cod_produs=produs["cod_produs"], 
                                cantitate_produs=produs["cantitate_produs"], 
                                pret_achizitie_produs=produs["pret_achizitie_produs"], 
                                moneda_achizitie = produs["moneda_achizitie"],
                                pret_vanzare_produs=produs["pret_vanzare_produs"],
                                moneda_vanzare = produs["moneda_vanzare"],
                                nir_id = nir_nou.id)
                    db.session.add(produs_nou)
                    db.session.commit()
                    produs_nou.pret_vanzare_produs_total = produs_nou.pret_vanzare_produs_total * produs_nou.cantitate_produs
                    db.session.commit()
                session.clear()
                session["data"] = []
                manipulareDbPentruGenerareNIR(nir_nou)
                return redirect(url_for('views.genereaza_nir'))
    else:
        return render_template("generateNIR.html", user=current_user, headings=headings, data=session.get("data"))

def manipulareDbPentruGenerareNIR(nir):
    # Get exchange rate for USD/RON on invoice date
    url = "https://exchange-rates.abstractapi.com/v1/historical/?api_key=30dc600e0cfd4e1fba1db37c8f521f42&base=USD&target=RON&date="+nir.data_factura
    response = requests.get(url)
    response = json.loads(response.content)
    response = response["exchange_rates"]["RON"]
    rata_schimb_valutar = response
    numar_produse=0
    for produs in nir.produse:
        numar_produse = numar_produse + produs.cantitate_produs
        if produs.moneda_achizitie == "USD" :
            produs.pret_achizitie_produs = produs.pret_achizitie_produs * rata_schimb_valutar
            produs.moneda_achizitie = "RON"
            db.session.commit()
    if nir.moneda_transport == "USD":
        nir.cost_transport = nir.cost_transport * rata_schimb_valutar
        nir.moneda_transport = "RON"
        db.session.commit()
    if nir.moneda_taxe_tva == "USD":
        nir.taxe_tva = nir.taxe_tva * rata_schimb_valutar
        nir.moneda_taxe_tva = "RON"
        db.session.commit()
    # Calcul adaos unitar pentru fiecare buc produs
    adaos_pret_achizitie_unitar = (nir.cost_transport + nir.taxe_tva)/numar_produse
    for produs in nir.produse:
        #Update pret achizitie produs buc cu adaosul rezultat din taxe si transport
        produs.pret_achizitie_produs = produs.pret_achizitie_produs + adaos_pret_achizitie_unitar
        produs.pret_achizitie_produs = round(produs.pret_achizitie_produs,2)
        produs.pret_achizitie_produs_total = produs.pret_achizitie_produs * produs.cantitate_produs
        produs.pret_vanzare_produs_total = produs.pret_vanzare_produs * produs.cantitate_produs
        #Update valoare adaos valoric per buc produs 
        produs.adaos_valoric_unitar = produs.pret_vanzare_produs - produs.pret_achizitie_produs
        #Updarte valoare adaos valoric total per produs
        produs.adaos_valoric_total = produs.adaos_valoric_unitar * produs.cantitate_produs
        #Update valoare adaos procentual per produs
        produs.adaos_procentual = ((produs.pret_vanzare_produs - produs.pret_achizitie_produs)/ produs.pret_achizitie_produs ) *100
        produs.adaos_procentual = round(produs.adaos_procentual,2)
        db.session.commit()

def styleTabel(ws, maxRow, coloana):
    fontStyle = Font(size = "10")
    ws[coloana + str(maxRow + 1)].alignment = Alignment(horizontal='center', vertical='center')
    ws[coloana + str(maxRow + 1)].border = Border(left=Side(border_style='thin', color='000000'),
                                                      right=Side(border_style='thin', color='000000'),
                                                      top=Side(border_style='thin', color='000000'),
                                                      bottom=Side(border_style='thin', color='000000'))
    ws[coloana + str(maxRow + 1)].font = fontStyle

@login_required
def genereazaExcel(nir):
    cwd  = os.getcwd()
    templateNir = load_workbook(cwd + "/website/static/templateNIR1.xlsx")
    ws = templateNir["TemplateNIR"]
    ws['C3'] = nir.nume_firma
    ws['I3'] = "NI-" + str(nir.id)
    ws['I4'] = nir.date
    ws['B6'] = "Subsemnatii, membrii ai comisiei de receptie, am receptionat valorile materiale furnizate de: "+ nir.nume_furnizor +" pe baza documentelor insotitoare: Factura NR: "+ nir.numar_factura + " ,constatand:"
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=14)
    nr_produse_diferite = len(nir.produse)
    ws.insert_rows(11, amount=nr_produse_diferite)
    totalAchizitie=0
    totalAdaos =0 
    totalVanzare =0
    nrTotalProduse = 0
    for i,produs in enumerate(nir.produse):
        ws.merge_cells(start_row=11 + i, start_column=2, end_row=11 + i, end_column=4)
        ws['A' + str(11 + i)] = i + 1
        ws['B' + str(11 + i)] = produs.nume_produs
        ws['E' + str(11 + i)] = "B.C"
        ws['F' + str(11 + i)] = produs.cantitate_produs
        ws['G' + str(11 + i)] = produs.cantitate_produs
        ws['H' + str(11 + i)] = produs.pret_achizitie_produs
        ws['I' + str(11 + i)] = produs.pret_achizitie_produs_total
        ws['J' + str(11 + i)] = produs.adaos_procentual
        ws['K' + str(11 + i)] = produs.adaos_valoric_unitar
        ws['L' + str(11 + i)] = produs.adaos_valoric_total
        ws['M' + str(11 + i)] = produs.pret_vanzare_produs
        ws['N' + str(11 + i)] = produs.pret_vanzare_produs_total
        styleTabel(ws, 11 + i, 'A')
        styleTabel(ws, 11 + i - 1, 'B')
        styleTabel(ws, 11 + i - 1, 'C')
        styleTabel(ws, 11 + i - 1, 'D')
        styleTabel(ws, 11 + i - 1, 'E')
        styleTabel(ws, 11 + i - 1, 'F')
        styleTabel(ws, 11 + i - 1, 'G')
        styleTabel(ws, 11 + i - 1, 'H')
        styleTabel(ws, 11 + i - 1, 'I')
        styleTabel(ws, 11 + i - 1, 'M')
        styleTabel(ws, 11 + i - 1, 'N')
        styleTabel(ws, 11 + i - 1, 'J')
        styleTabel(ws, 11 + i - 1, 'K')
        styleTabel(ws, 11 + i - 1, 'L')
        totalAchizitie = totalAchizitie + produs.pret_achizitie_produs_total
        totalAdaos = totalAdaos + produs.adaos_valoric_total
        totalVanzare = totalVanzare + produs.pret_vanzare_produs_total
        nrTotalProduse = nrTotalProduse + produs.cantitate_produs
    ws['G' + str(11 + nr_produse_diferite)] = nrTotalProduse
    ws['H' + str(11 + nr_produse_diferite)] = totalAchizitie
    ws['J' + str(11 + nr_produse_diferite)] = totalAdaos
    ws['M' + str(11 + nr_produse_diferite)] = totalVanzare
    ws['B' + str(16+ nr_produse_diferite)] = nir.nume_membri
    ws['K' + str(16+ nr_produse_diferite)] = nir.nume_gestionar
    ws.merge_cells(start_row=int(11 + nr_produse_diferite), start_column=8, end_row=int(11 + nr_produse_diferite), end_column=9)
    ws.merge_cells(start_row=int(11 + nr_produse_diferite), start_column=10, end_row=int(11 + nr_produse_diferite), end_column=12)
    ws.merge_cells(start_row=int(11 + nr_produse_diferite), start_column=13, end_row=int(11 + nr_produse_diferite), end_column=14)
    NirPath = cwd + "/website/static/nirs/" + str(current_user.id) + "/nir-"+str(nir.id)+".xlsx"
    templateNir.save(NirPath)

@views.route('/sterge-produs', methods=['POST'])
def sterge_produs():
    produs = json.loads(request.data)
    cod_produs = produs['button_id']
    for produs in session.get("data"):
        if produs["cod_produs"] == cod_produs:
            session.get('data').remove(produs)
            flash('Produs sters', category='succcess')
    return jsonify({})

@views.route('/sterge-nir', methods=['POST'])
def sterge_nir():
    nir = json.loads(request.data)
    nir_id = nir['nir_id']
    nir = Nir.query.get(nir_id)
    if nir:
        if nir.user_id == current_user.id:
            db.session.delete(nir)
            db.session.commit()
    return jsonify({})

@views.route('/get-nir/nir-<nir_id>.xlsx')
@login_required
def get_nir(nir_id):
    path = os.getcwd()
    path = path +"/website/static/nirs/"+ str(current_user.id)
    nir_name="nir-"+nir_id+".xlsx"
    try:
        nir = Nir.query.get(nir_id)
        if nir:
            if nir.user_id == current_user.id:
                genereazaExcel(nir)
                return send_from_directory(path, path=nir_name, as_attachment=True) 
    except FileNotFoundError:
        abort(404)